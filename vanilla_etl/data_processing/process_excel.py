from openpyxl import load_workbook
import pandas as pd
from prefect import flow, task, get_run_logger
from prefect.task_runners import ConcurrentTaskRunner
from vanilla_etl.data_processing.tables import identify_tables
from typing import List
import numpy as np
from deep_translator import GoogleTranslator


@task
def load_excel_file(file_path: str, config) -> List[pd.DataFrame]:
    """
    Process a single Excel file, extract potential tables from each sheet.

    Args:
        file_path (str): Path to the Excel file.
        config (dict): Configuration for identifying tables.

    Returns:
        List[pd.DataFrame]: List of DataFrames containing extracted tables.
    """
    logger = get_run_logger()
    tables = []

    try:
        # Load workbook and process each sheet
        workbook = load_workbook(file_path, data_only=True)
        for sheet_name in workbook.sheetnames:
            logger.info(f"Processing sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            data = [[cell.value for cell in row] for row in sheet.iter_rows()]
            df = pd.DataFrame(data)

            # Identify and extract potential tables from the sheet
            tables_in_sheet = identify_tables(df, config)
            tables.extend(tables_in_sheet)

            if not tables:
                raise ValueError("No tables found in the workbook.")

            logger.info(f"Extracted {len(tables)} tables from file: {file_path}")
            return tables

    except Exception as e:
        logger.error(f"Error processing file '{file_path}': {e}")
        raise


@flow(task_runner=ConcurrentTaskRunner())
def load_data_concurrently(file_paths: List[str], config) -> List[List[pd.DataFrame]]:
    """
    Load and process multiple Excel files concurrently.

    Args:
        file_paths (List[str]): List of file paths to Excel files.
        config (dict): Configuration for identifying tables.

    Returns:
        List[List[pd.DataFrame]]: List of DataFrames extracted from each file.
    """
    dataframes = [load_excel_file.submit(file_path, config) for file_path in file_paths]
    return dataframes
